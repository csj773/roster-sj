from __future__ import annotations

import argparse
import base64
import json
import os
import re
import shutil
import smtplib
import subprocess
import time
from datetime import datetime
from email.message import EmailMessage
from pathlib import Path
from typing import Any

from google.oauth2 import service_account
from googleapiclient.errors import HttpError
from googleapiclient.discovery import build
from openpyxl import Workbook, load_workbook


ROOT = Path(os.environ.get("LOGBOOK_ROOT", Path(__file__).resolve().parents[1])).resolve()
WORK = ROOT / "work"
OUT = ROOT / "outputs"
SUMMARY_JSON = WORK / "github_automation_summary.json"

PILOTLOG_SPREADSHEET_ID = os.environ.get("PILOTLOG_SPREADSHEET_ID", "1mKjEd__zIoMJaa6CLmDE-wALGhtlG-USLTAiQBZnioc")
AUTHORITATIVE_SPREADSHEET_ID = os.environ.get("AUTHORITATIVE_SPREADSHEET_ID", "1ZasXUvOP8jl8lpNZVFp5CENFNZc-o1sIiWURJ2mSmWg")
SHEET_NAME = os.environ.get("LOGBOOK_SHEET_NAME", "flt_log")
CHECKPOINT_SHEET_NAME = os.environ.get("PILOTLOG_CHECKPOINT_SHEET_NAME", "_pilotlog_checkpoint")
CHECKPOINT_FILE = WORK / "pilotlog_checkpoint_rows.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
]
ILLEGAL_XLSX_CHARS = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")
RETRYABLE_GOOGLE_STATUSES = {429, 500, 502, 503, 504}


def require_env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing required environment variable: {name}")
    return value


def service_account_info() -> dict[str, Any]:
    raw = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON", "").strip()
    b64 = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON_B64", "").strip()
    if b64:
        raw = base64.b64decode(b64).decode("utf-8")
    if not raw:
        raise RuntimeError("Set GOOGLE_SERVICE_ACCOUNT_JSON or GOOGLE_SERVICE_ACCOUNT_JSON_B64.")
    return json.loads(raw)


def google_clients() -> tuple[Any, str]:
    info = service_account_info()
    client_email = info.get("client_email", "<missing client_email>")
    print(f"Using Google service account: {client_email}")
    creds = service_account.Credentials.from_service_account_info(info, scopes=SCOPES)
    return build("sheets", "v4", credentials=creds), client_email


def execute_google_request(request: Any, *, attempts: int = 5) -> Any:
    for attempt in range(1, attempts + 1):
        try:
            return request.execute()
        except HttpError as exc:
            status = getattr(exc.resp, "status", None)
            if status not in RETRYABLE_GOOGLE_STATUSES or attempt == attempts:
                raise
            delay = min(60, 2 ** attempt)
            print(f"Google API returned {status}; retrying in {delay}s ({attempt}/{attempts})")
            time.sleep(delay)
    raise RuntimeError("Google API request failed without returning a response.")


def export_sheet_values_xlsx(sheets: Any, spreadsheet_id: str, destination: Path, client_email: str, *, allow_rename: bool) -> None:
    if allow_rename:
        ensure_sheet_name(sheets, spreadsheet_id, client_email)
    else:
        require_sheet_exists(sheets, spreadsheet_id, client_email)
    result = sheets.spreadsheets().values().get(
        spreadsheetId=spreadsheet_id,
        range=f"{SHEET_NAME}!A:AZ",
        valueRenderOption="FORMATTED_VALUE",
        dateTimeRenderOption="FORMATTED_STRING",
    )
    result = execute_google_request(result)
    values = result.get("values", [])
    if not values:
        raise RuntimeError(f"No values returned from spreadsheet {spreadsheet_id} sheet {SHEET_NAME}.")
    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    for row in values:
        ws.append([clean_xlsx_value(value) for value in row])
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)
    print(f"Wrote {destination} from Sheets API values: {len(values)} rows")


def clean_xlsx_value(value: Any) -> Any:
    if isinstance(value, str):
        return ILLEGAL_XLSX_CHARS.sub("", value)
    return value


def ensure_sheet_name(sheets: Any, spreadsheet_id: str, client_email: str) -> int:
    try:
        request = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties")
        metadata = execute_google_request(request)
    except HttpError as exc:
        if exc.resp.status == 403:
            raise RuntimeError(
                "Google Sheets permission denied. Share spreadsheet "
                f"{spreadsheet_id} with service account {client_email} as Editor, "
                "then rerun the workflow."
            ) from exc
        raise
    sheet_props = [sheet["properties"] for sheet in metadata.get("sheets", []) if sheet["properties"].get("sheetType") == "GRID"]
    for props in sheet_props:
        if props.get("title") == SHEET_NAME:
            return int(props["sheetId"])
    if len(sheet_props) != 1:
        titles = ", ".join(props.get("title", "") for props in sheet_props)
        raise RuntimeError(f"Authoritative spreadsheet has no {SHEET_NAME!r} tab and is not single-sheet. Tabs: {titles}")
    sheet_id = int(sheet_props[0]["sheetId"])
    request = sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={"requests": [{"updateSheetProperties": {"properties": {"sheetId": sheet_id, "title": SHEET_NAME}, "fields": "title"}}]},
    )
    execute_google_request(request)
    return sheet_id


def require_sheet_exists(sheets: Any, spreadsheet_id: str, client_email: str) -> int:
    try:
        request = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties")
        metadata = execute_google_request(request)
    except HttpError as exc:
        if exc.resp.status == 403:
            raise RuntimeError(
                "Google Sheets permission denied. Share spreadsheet "
                f"{spreadsheet_id} with service account {client_email} as Editor, "
                "then rerun the workflow."
            ) from exc
        raise
    for sheet in metadata.get("sheets", []):
        props = sheet["properties"]
        if props.get("sheetType") == "GRID" and props.get("title") == SHEET_NAME:
            return int(props["sheetId"])
    titles = ", ".join(sheet["properties"].get("title", "") for sheet in metadata.get("sheets", []))
    raise RuntimeError(f"Spreadsheet {spreadsheet_id} does not contain required sheet {SHEET_NAME!r}. Tabs: {titles}")


def ensure_checkpoint_sheet(sheets: Any, spreadsheet_id: str, client_email: str) -> int:
    try:
        request = sheets.spreadsheets().get(spreadsheetId=spreadsheet_id, fields="sheets.properties")
        metadata = execute_google_request(request)
    except HttpError as exc:
        if exc.resp.status == 403:
            raise RuntimeError(
                "Google Sheets permission denied. Share spreadsheet "
                f"{spreadsheet_id} with service account {client_email} as Editor, "
                "then rerun the workflow."
            ) from exc
        raise

    for sheet in metadata.get("sheets", []):
        props = sheet["properties"]
        if props.get("sheetType") == "GRID" and props.get("title") == CHECKPOINT_SHEET_NAME:
            sheet_id = int(props["sheetId"])
            if not props.get("hidden"):
                request = sheets.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        "requests": [
                            {
                                "updateSheetProperties": {
                                    "properties": {"sheetId": sheet_id, "hidden": True},
                                    "fields": "hidden",
                                }
                            }
                        ]
                    },
                )
                execute_google_request(request)
            return sheet_id

    request = sheets.spreadsheets().batchUpdate(
        spreadsheetId=spreadsheet_id,
        body={
            "requests": [
                {
                    "addSheet": {
                        "properties": {
                            "title": CHECKPOINT_SHEET_NAME,
                            "hidden": True,
                            "gridProperties": {"rowCount": 1000, "columnCount": 2},
                        }
                    }
                }
            ]
        },
    )
    response = execute_google_request(request)
    return int(response["replies"][0]["addSheet"]["properties"]["sheetId"])


def download_checkpoint(sheets: Any, client_email: str) -> None:
    ensure_checkpoint_sheet(sheets, AUTHORITATIVE_SPREADSHEET_ID, client_email)
    request = sheets.spreadsheets().values().get(
        spreadsheetId=AUTHORITATIVE_SPREADSHEET_ID,
        range=f"{CHECKPOINT_SHEET_NAME}!A:B",
    )
    response = execute_google_request(request)
    values = response.get("values", [])
    CHECKPOINT_FILE.write_text(json.dumps(values, ensure_ascii=False), encoding="utf-8")


def upload_checkpoint(sheets: Any, client_email: str) -> None:
    if not CHECKPOINT_FILE.exists():
        return
    ensure_checkpoint_sheet(sheets, AUTHORITATIVE_SPREADSHEET_ID, client_email)
    values = json.loads(CHECKPOINT_FILE.read_text(encoding="utf-8"))
    request = sheets.spreadsheets().values().clear(
        spreadsheetId=AUTHORITATIVE_SPREADSHEET_ID,
        range=f"{CHECKPOINT_SHEET_NAME}!A:B",
        body={},
    )
    execute_google_request(request)
    if values:
        request = sheets.spreadsheets().values().update(
            spreadsheetId=AUTHORITATIVE_SPREADSHEET_ID,
            range=f"{CHECKPOINT_SHEET_NAME}!A1",
            valueInputOption="RAW",
            body={"values": values},
        )
        execute_google_request(request)


def cell_to_sheets_value(cell: Any) -> str | int | float:
    value = cell.value
    if value is None:
        return ""
    if cell.data_type == "f":
        formula = f"={value}" if not str(value).startswith("=") else str(value)
        return formula.replace("_xlfn.IFS", "IFS")
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if hasattr(value, "isoformat") and value.__class__.__name__ == "time":
        return value.strftime("%H:%M")
    if hasattr(value, "total_seconds"):
        mins = int(round(value.total_seconds() / 60))
        return f"{mins // 60}:{mins % 60:02d}"
    return value


def update_authoritative_sheet(sheets: Any, synced_xlsx: Path, client_email: str) -> None:
    ensure_sheet_name(sheets, AUTHORITATIVE_SPREADSHEET_ID, client_email)
    wb = load_workbook(synced_xlsx, data_only=False)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    values = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        values.append([cell_to_sheets_value(cell) for cell in row])
    request = sheets.spreadsheets().values().clear(
        spreadsheetId=AUTHORITATIVE_SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A:Z",
        body={},
    )
    execute_google_request(request)
    request = sheets.spreadsheets().values().update(
        spreadsheetId=AUTHORITATIVE_SPREADSHEET_ID,
        range=f"{SHEET_NAME}!A1",
        valueInputOption="USER_ENTERED",
        body={"values": values},
    )
    execute_google_request(request)


def run_python(script: str, env_overrides: dict[str, str] | None = None) -> str:
    env = {**os.environ, "LOGBOOK_ROOT": str(ROOT)}
    if env_overrides:
        env.update(env_overrides)
    result = subprocess.run(
        ["python", str(WORK / script)],
        cwd=str(ROOT),
        env=env,
        text=True,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
    )
    print(result.stdout)
    result.check_returncode()
    return result.stdout


def parse_lines(text: str) -> dict[str, str]:
    values: dict[str, str] = {}
    for line in text.splitlines():
        parts = line.split(maxsplit=1)
        if len(parts) == 2:
            values[parts[0]] = parts[1]
    return values


def prefixed(values: dict[str, str], prefix: str) -> dict[str, str]:
    return {f"{prefix}_{key}": value for key, value in values.items()}


def logbook_attachments() -> list[Path]:
    preferred = [
        OUT / "log filled V1.xlsx",
        OUT / "ICAO_EASA_A4_landscape_logbook_V1.pdf",
        OUT / "A5_booklet_A4_portrait_duplex_V1.pdf",
        OUT / "log filled V2.xlsx",
        OUT / "ICAO_EASA_A4_landscape_logbook_V2.pdf",
        OUT / "A5_booklet_A4_portrait_duplex_V2.pdf",
        OUT / "B5_print_logbook_V2.pdf",
    ]
    if all(path.exists() for path in preferred[:3]):
        return [path for path in preferred if path.exists()]
    return [
        OUT / "log filled.xlsx",
        OUT / "ICAO_EASA_A4_landscape_logbook.pdf",
        OUT / "A5_booklet_A4_portrait_duplex.pdf",
    ]


def send_email(summary: dict[str, str]) -> None:
    if os.environ.get("SEND_EMAIL", "true").lower() in {"0", "false", "no"}:
        print("SEND_EMAIL disabled; skipping SMTP delivery.")
        return
    sender = require_env("GMAIL_USERNAME")
    password = require_env("GMAIL_APP_PASSWORD")
    recipient = os.environ.get("LOGBOOK_EMAIL_TO", "sjchoi787@gmail.com")
    today = os.environ.get("RUN_DATE", datetime.now().strftime("%Y-%m-%d"))[:10]
    attachments = logbook_attachments()
    for attachment in attachments:
        if not attachment.exists():
            raise RuntimeError(f"Missing attachment: {attachment}")

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = recipient
    msg["Subject"] = f"PILOTLOG Logbook Update - {today}"
    msg.set_content(
        "\n".join(
            [
                f"PILOTLOG logbook update completed on {today}.",
                "",
                f"Added rows: {summary.get('SYNC_ADDED', summary.get('ADDED', 'n/a'))}",
                f"Deleted rows: {summary.get('SYNC_DELETED', summary.get('DELETED', 'n/a'))}",
                f"Modified rows: {summary.get('SYNC_MODIFIED', summary.get('MODIFIED', 'n/a'))}",
                f"V1 flight count: {summary.get('V1_FINAL_FLIGHTS', summary.get('FINAL_FLIGHTS', 'n/a'))}",
                f"V1 cumulative totals: {summary.get('V1_FINAL_TOTALS', summary.get('FINAL_TOTALS', 'n/a'))}",
                f"V2 flight count: {summary.get('V2_FINAL_FLIGHTS', 'n/a')}",
                f"V2 cumulative totals: {summary.get('V2_FINAL_TOTALS', 'n/a')}",
                f"Formula errors: V1 {summary.get('V1_FORMULA_ERRORS', summary.get('FORMULA_ERRORS', 'n/a'))}, V2 {summary.get('V2_FORMULA_ERRORS', 'n/a')}",
                "",
                "Attached files: " + ", ".join(path.name for path in attachments),
                "Print note: print the A5 booklet PDF on A4 portrait paper, duplex, short-edge flip. V2 also includes a B5 print PDF with cover.",
            ]
        )
    )
    for attachment in attachments:
        msg.add_attachment(
            attachment.read_bytes(),
            maintype="application",
            subtype="octet-stream",
            filename=attachment.name,
        )
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(sender, password)
        smtp.send_message(msg)


def build_logbook() -> dict[str, str]:
    WORK.mkdir(parents=True, exist_ok=True)
    OUT.mkdir(parents=True, exist_ok=True)
    sheets, client_email = google_clients()

    ensure_sheet_name(sheets, AUTHORITATIVE_SPREADSHEET_ID, client_email)
    download_checkpoint(sheets, client_email)
    authoritative_download = WORK / "log_filled_authoritative_download.xlsx"
    export_sheet_values_xlsx(sheets, AUTHORITATIVE_SPREADSHEET_ID, authoritative_download, client_email, allow_rename=True)
    export_sheet_values_xlsx(sheets, PILOTLOG_SPREADSHEET_ID, WORK / "PILOTLOG_export.xlsx", client_email, allow_rename=False)

    previous_output = OUT / "log filled.xlsx"
    if not previous_output.exists():
        shutil.copy2(authoritative_download, previous_output)
    shutil.copy2(authoritative_download, WORK / "log_filled_authoritative_synced.xlsx")

    sync_output = run_python("sync_authoritative_from_pilotlog.py")
    update_authoritative_sheet(sheets, WORK / "log_filled_authoritative_synced.xlsx", client_email)
    upload_checkpoint(sheets, client_email)

    v1_output = run_python("build_final_deliverables.py")
    shutil.copy2(OUT / "log filled.xlsx", OUT / "log filled V1.xlsx")
    shutil.copy2(OUT / "ICAO_EASA_A4_landscape_logbook.pdf", OUT / "ICAO_EASA_A4_landscape_logbook_V1.pdf")
    shutil.copy2(OUT / "A5_booklet_A4_portrait_duplex.pdf", OUT / "A5_booklet_A4_portrait_duplex_V1.pdf")

    v2_output = run_python(
        "build_final_deliverables.py",
        {
            "LOGBOOK_SOURCE": "pilotlog",
            "LOGBOOK_START_DATE": os.environ.get("LOGBOOK_V2_START_DATE", "2026-06-26"),
            "LOGBOOK_END_DATE": "",
            "LOGBOOK_EXPECTED_FINAL_BLK": "",
            "LOGBOOK_MIN_FLIGHTS": "0",
            "LOGBOOK_ALLOW_EMPTY": "true",
            "LOGBOOK_START_BLK": os.environ.get("LOGBOOK_V2_START_BLK", "16129:30"),
            "LOGBOOK_START_NGT": os.environ.get("LOGBOOK_V2_START_NGT", "5942:04"),
            "LOGBOOK_START_IFR": os.environ.get("LOGBOOK_V2_START_IFR", "11927:45"),
            "LOGBOOK_START_TO": os.environ.get("LOGBOOK_V2_START_TO", "1827"),
            "LOGBOOK_START_LDG": os.environ.get("LOGBOOK_V2_START_LDG", "1836"),
            "LOGBOOK_START_PIC": os.environ.get("LOGBOOK_V2_START_PIC", "7638:40"),
            "LOGBOOK_XLSX_OUT": str(OUT / "log filled V2.xlsx"),
            "LOGBOOK_A4_OUT": str(OUT / "ICAO_EASA_A4_landscape_logbook_V2.pdf"),
            "LOGBOOK_BOOKLET_OUT": str(OUT / "A5_booklet_A4_portrait_duplex_V2.pdf"),
            "LOGBOOK_B5_OUT": str(OUT / "B5_print_logbook_V2.pdf"),
        },
    )
    summary = {
        **parse_lines(sync_output),
        **prefixed(parse_lines(v1_output), "V1"),
        **prefixed(parse_lines(v2_output), "V2"),
    }
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"SUMMARY_JSON {SUMMARY_JSON}")
    return summary


def read_summary() -> dict[str, str]:
    if not SUMMARY_JSON.exists():
        return {}
    return json.loads(SUMMARY_JSON.read_text(encoding="utf-8"))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--send-email-only", action="store_true", help="Send already generated deliverables without rebuilding.")
    parser.add_argument("--skip-email", action="store_true", help="Build deliverables and skip SMTP delivery.")
    args = parser.parse_args()
    if args.send_email_only:
        send_email(read_summary())
        return
    summary = build_logbook()
    if args.skip_email:
        print("Email delivery skipped by --skip-email.")
        return
    send_email(summary)


if __name__ == "__main__":
    main()
